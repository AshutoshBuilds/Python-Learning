"""
Lesson 30: Excel with openpyxl
==============================
NOTE: pip install openpyxl

Write a GST summary workbook and read it back — mirrors real
accounting deliverables (GST reports for filing / audit).

Run: python 30_excel_openpyxl.py
"""

import sys
from pathlib import Path

# =============================================================================
# CONCEPT
# =============================================================================
# openpyxl reads and writes Excel .xlsx files from Python.
#
#   Workbook()           — create new workbook
#   ws.append([...])     — add row
#   wb.save(path)        — save file
#   load_workbook(path)  — open existing file
#
# Commerce use:
#   GST summary sheets, trial balance exports, invoice registers

print("=" * 60)
print("LESSON 30: Excel with openpyxl")
print("=" * 60)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("\nERROR: openpyxl is not installed.")
    print("Install with: pip install openpyxl")
    print("Then re-run this lesson.")
    sys.exit(0)

# =============================================================================
# LIVE DEMOS
# =============================================================================

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)
OUTPUT_PATH = DATA_DIR / "gst_summary_demo.xlsx"

# Sample GST data (as if aggregated from sales register)
gst_records = [
    {"gst_rate": "5%",  "taxable_value": 10700.0,  "cgst": 267.50,  "sgst": 267.50,  "igst": 0.0},
    {"gst_rate": "12%", "taxable_value": 5425.0,   "cgst": 325.50,  "sgst": 325.50,  "igst": 0.0},
    {"gst_rate": "18%", "taxable_value": 18970.0,  "cgst": 1707.30, "sgst": 1707.30, "igst": 0.0},
]


def write_gst_summary(path, records, business_name="Demo Commerce Pvt Ltd"):
    """Write GST summary sheet to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "GST Summary"

    # Header
    ws["A1"] = "GST Summary Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = business_name
    ws["A3"] = "Period: Jan 2026"

    # Column headers (row 5)
    headers = ["GST Rate", "Taxable Value", "CGST", "SGST", "IGST", "Total GST"]
    ws.append([])
    ws.append([])
    ws.append(headers)
    for cell in ws[5]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for rec in records:
        total_gst = rec["cgst"] + rec["sgst"] + rec["igst"]
        ws.append([
            rec["gst_rate"],
            rec["taxable_value"],
            rec["cgst"],
            rec["sgst"],
            rec["igst"],
            total_gst,
        ])

    # Totals row
    total_taxable = sum(r["taxable_value"] for r in records)
    total_cgst = sum(r["cgst"] for r in records)
    total_sgst = sum(r["sgst"] for r in records)
    total_igst = sum(r["igst"] for r in records)
    total_gst_all = total_cgst + total_sgst + total_igst

    ws.append([])
    ws.append(["TOTAL", total_taxable, total_cgst, total_sgst, total_igst, total_gst_all])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 12
    for col in "BCDEF":
        ws.column_dimensions[col].width = 16

    wb.save(path)
    return total_taxable, total_gst_all


def read_gst_summary(path):
    """Read back GST summary and return list of data rows."""
    wb = load_workbook(path, data_only=True)
    ws = wb["GST Summary"]
    rows = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if row[0] is None or row[0] == "TOTAL":
            continue
        rows.append({
            "gst_rate": row[0],
            "taxable_value": float(row[1]),
            "total_gst": float(row[5]),
        })
    return rows


# --- Demo: Write Excel ---
print("\n--- Writing GST Summary Excel ---")
total_taxable, total_gst = write_gst_summary(OUTPUT_PATH, gst_records)
print(f"Created: {OUTPUT_PATH}")
print(f"  Total taxable: Rs {total_taxable:,.2f}")
print(f"  Total GST:     Rs {total_gst:,.2f}")

# --- Demo: Read Excel back ---
print("\n--- Reading GST Summary Back ---")
loaded = read_gst_summary(OUTPUT_PATH)
print(f"{'Rate':<8} {'Taxable':>14} {'Total GST':>14}")
print("-" * 38)
for row in loaded:
    print(f"{row['gst_rate']:<8} {row['taxable_value']:>14,.2f} {row['total_gst']:>14,.2f}")

# --- Demo: Inspect workbook structure ---
print("\n--- Workbook Structure ---")
wb = load_workbook(OUTPUT_PATH)
print(f"Sheet names: {wb.sheetnames}")
ws = wb.active
print(f"Active sheet: {ws.title}")
print(f"Used rows: {ws.max_row}, columns: {ws.max_column}")
print(f"Title cell A1: {ws['A1'].value}")


# =============================================================================
# EXERCISES
# =============================================================================
print("\n" + "=" * 60)
print("EXERCISES (try yourself, then uncomment solutions)")
print("=" * 60)

# Exercise 1: Open gst_summary_demo.xlsx and print cell B6 (first taxable value).

# wb = load_workbook(OUTPUT_PATH, data_only=True)
# print(wb["GST Summary"]["B6"].value)

# Exercise 2: Add a second sheet "Notes" with one row: "Generated for training".

# wb = load_workbook(OUTPUT_PATH)
# notes = wb.create_sheet("Notes")
# notes["A1"] = "Generated for training"
# wb.save(OUTPUT_PATH)

# Exercise 3: Count data rows (exclude header and TOTAL) programmatically.

# rows = read_gst_summary(OUTPUT_PATH)
# print(f"Data rows: {len(rows)}")


# =============================================================================
# MINI CHALLENGE
# =============================================================================
print("\n" + "=" * 60)
print("MINI CHALLENGE")
print("=" * 60)
print("""
1. Create a new file data/gst_summary_challenge.xlsx
2. Add one row: 28% GST (luxury demo), taxable 10000, split CGST/SGST equally
3. Read it back and verify total GST = 2800
""")

# --- Mini challenge reference solution ---
challenge_records = [
    {"gst_rate": "28%", "taxable_value": 10000.0, "cgst": 1400.0, "sgst": 1400.0, "igst": 0.0},
]
challenge_path = DATA_DIR / "gst_summary_challenge.xlsx"
_, challenge_gst = write_gst_summary(challenge_path, challenge_records, business_name="Challenge Co")
read_back = read_gst_summary(challenge_path)

print("\nChallenge result:")
print(f"  File: {challenge_path.name}")
print(f"  GST written: Rs {challenge_gst:,.2f}")
print(f"  GST read back: Rs {read_back[0]['total_gst']:,.2f}")
print(f"  Verified: {challenge_gst == 2800.0}")

print("\nLesson 30 complete.")
